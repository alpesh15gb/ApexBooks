"""
Export engine: XLSX, CSV, PDF report generation.

Supports streaming for large datasets and enterprise formatting.
"""

import csv
import io
import json
import os
import uuid
from datetime import datetime
from typing import Any, Optional
from app.core.exceptions import APIError


try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    openpyxl = None


class ExportFormat:
    XLSX = 'xlsx'
    CSV = 'csv'
    JSON = 'json'


if HAS_OPENPYXL:
    STYLING = {
        'header_fill': PatternFill(start_color='1F2937', end_color='1F2937', fill_type='solid'),
        'header_font': Font(name='Calibri', bold=True, size=11, color='FFFFFF'),
        'odd_row_fill': PatternFill(start_color='F9FAFB', end_color='F9FAFB', fill_type='solid'),
        'even_row_fill': PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid'),
        'data_font': Font(name='Calibri', size=10),
        'bold_font': Font(name='Calibri', bold=True, size=10),
        'title_font': Font(name='Calibri', bold=True, size=14, color='1F2937'),
        'subtitle_font': Font(name='Calibri', size=10, color='6B7280'),
        'money_format': '#,##0.00',
        'date_format': 'DD-MMM-YYYY',
        'thin_border': Border(
            left=Side(style='thin', color='E5E7EB'),
            right=Side(style='thin', color='E5E7EB'),
            top=Side(style='thin', color='E5E7EB'),
            bottom=Side(style='thin', color='E5E7EB'),
        ),
    }
else:
    STYLING = {}


class ReportExporter:
    """Export report data to XLSX, CSV, JSON with enterprise formatting."""

    def __init__(self, report_data: dict, report_name: str = "Report"):
        self.data = report_data
        self.report_name = report_name
        self.timestamp = datetime.now()

    # ── XLSX Export ────────────────────────────────────────────────

    def to_xlsx(self) -> bytes:
        """Generate a professionally formatted XLSX workbook."""
        if not HAS_OPENPYXL:
            return self._fallback_xlsx()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = self.report_name[:31]

        # Extract rows from different report structures
        rows = self._extract_rows()

        # Title row
        ws.merge_cells('A1:H1')
        ws['A1'] = self.report_name
        ws['A1'].font = STYLING['title_font']
        ws['A1'].alignment = Alignment(horizontal='left')

        # Subtitle row
        ws.merge_cells('A2:H2')
        period = self.data.get('period') or self.data.get('as_of_date') or ''
        ws['A2'] = f"Generated: {self.timestamp.strftime('%d-%b-%Y %H:%M')} | Period: {period}"
        ws['A2'].font = STYLING['subtitle_font']

        # Headers in row 4
        if rows:
            headers = list(rows[0].keys())
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col_idx, value=header.replace('_', ' ').title())
                cell.font = STYLING['header_font']
                cell.fill = STYLING['header_fill']
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = STYLING['thin_border']

            # Data rows starting from row 5
            for row_idx, row_data in enumerate(rows, 5):
                fill = STYLING['odd_row_fill'] if (row_idx - 5) % 2 == 0 else STYLING['even_row_fill']
                for col_idx, header in enumerate(headers, 1):
                    value = row_data.get(header, '')
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.font = STYLING['data_font']
                    cell.fill = fill
                    cell.border = STYLING['thin_border']
                    cell.alignment = Alignment(horizontal='right' if isinstance(value, (int, float)) else 'left')

                    # Format money values
                    if isinstance(value, float) and any(h in header.lower() for h in ['amount', 'total', 'balance', 'debit', 'credit', 'tax', 'value', 'price', 'gst']):
                        cell.number_format = STYLING['money_format']

            # Auto-width columns
            for col_idx in range(1, len(headers) + 1):
                max_len = len(str(headers[col_idx - 1]))
                for row_idx in range(5, min(5 + len(rows), 105)):  # Sample first 100 rows for performance
                    cell_val = ws.cell(row=row_idx, column=col_idx).value
                    if cell_val:
                        max_len = max(max_len, len(str(cell_val)))
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 50)

            # Freeze panes (header row + title)
            ws.freeze_panes = 'A5'

            # Auto-filter
            if rows:
                last_col = get_column_letter(len(headers))
                ws.auto_filter.ref = f'A4:{last_col}{4 + len(rows)}'

        # Add summary sheet if available
        self._add_summary_sheet(wb)

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    def _add_summary_sheet(self, wb):
        """Add a summary/metadata sheet."""
        ws = wb.create_sheet('Summary')
        metadata = [
            ('Report', self.report_name),
            ('Generated', self.timestamp.isoformat()),
            ('Period', self.data.get('period', '')),
            ('Metrics', json.dumps(self.data.get('metrics', {}), default=str)),
        ]
        for i, (k, v) in enumerate(metadata, 1):
            ws.cell(row=i, column=1, value=k).font = STYLING['bold_font']
            ws.cell(row=i, column=2, value=v).font = STYLING['data_font']
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 60

    def _fallback_xlsx(self) -> bytes:
        """Fallback simple XLSX if openpyxl not available."""
        import zipfile
        # Minimal XLSX writer using XML
        rows = self._extract_rows()

        xml_parts = []
        # Shared strings
        ss = []
        for r in rows:
            for v in r.values():
                if isinstance(v, str) and v not in ss:
                    ss.append(v)

        # Build minimal workbook
        output = io.BytesIO()
        with zipfile.ZipFile(output, 'w') as zf:
            zf.writestr('[Content_Types].xml', self._xlsx_content_types())
            zf.writestr('_rels/.rels', self._xlsx_rels())
            zf.writestr('xl/workbook.xml', self._xlsx_workbook())
            zf.writestr('xl/_rels/workbook.xml.rels', self._xlsx_wb_rels())
            zf.writestr('xl/worksheets/sheet1.xml', self._xlsx_sheet(rows))
            zf.writestr('xl/sharedStrings.xml', self._xlsx_shared_strings(ss))
        return output.getvalue()

    def _xlsx_content_types(self): return '<?xml version="1.0"?><Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types"><Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/><Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/><Override PartName="/xl/sharedStrings.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sharedStrings+xml"/></Types>'
    def _xlsx_rels(self): return '<?xml version="1.0"?><Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"><Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/></Relationships>'
    def _xlsx_workbook(self): return '<?xml version="1.0"?><workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"><sheets><sheet name="Report" sheetId="1" r:id="rId1" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"/></sheets></workbook>'
    def _xlsx_wb_rels(self): return '<?xml version="1.0"?><Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"><Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/></Relationships>'
    def _xlsx_shared_strings(self, ss):
        items = ''.join(f'<si><t xml:space="preserve">{s}</t></si>' for s in ss)
        return f'<?xml version="1.0"?><sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" count="{len(ss)}" uniqueCount="{len(ss)}">{items}</sst>'
    def _xlsx_sheet(self, rows):
        if not rows:
            return '<?xml version="1.0"?><worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"><sheetData><row r="1"><c t="inlineStr" r="A1"><is><t>No data</t></is></c></row></sheetData></worksheet>'
        headers = list(rows[0].keys())
        col_letters = [get_column_letter(i + 1) for i in range(len(headers))]
        xml_rows = []
        # Header row
        hcells = ''.join(f'<c t="inlineStr" r="{col}{1}"><is><t>{h.replace("_"," ").title()}</t></is></c>' for col, h in zip(col_letters, headers))
        xml_rows.append(f'<row r="1">{hcells}</row>')
        # Data rows
        for ri, row in enumerate(rows, 2):
            cells = []
            for col, h in zip(col_letters, headers):
                v = row.get(h, '')
                if isinstance(v, (int, float)):
                    cells.append(f'<c r="{col}{ri}"><v>{v}</v></c>')
                else:
                    cells.append(f'<c t="inlineStr" r="{col}{ri}"><is><t>{str(v).replace("&","&amp;").replace("<","&lt;")}</t></is></c>')
            xml_rows.append(f'<row r="{ri}">{"".join(cells)}</row>')
        return f'<?xml version="1.0"?><worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"><sheetData>{"".join(xml_rows)}</sheetData></worksheet>'

    # ── CSV Export ─────────────────────────────────────────────────

    def to_csv(self, delimiter: str = ',') -> bytes:
        """Generate CSV with proper quoting."""
        output = io.StringIO()
        writer = csv.writer(output, delimiter=delimiter, quoting=csv.QUOTE_ALL)
        rows = self._extract_rows()
        if rows:
            writer.writerow(rows[0].keys())
            money_keys = {k for k in rows[0] if isinstance(rows[0][k], float)}
            for row in rows:
                writer.writerow([round(v, 2) if k in money_keys else v for k, v in row.items()])
        return output.getvalue().encode('utf-8-sig')

    # ── JSON Export ────────────────────────────────────────────────

    def to_json(self, indent: int = 2) -> bytes:
        """Generate JSON export."""
        return json.dumps(self.data, default=str, indent=indent).encode('utf-8')

    # ── Row Extraction ─────────────────────────────────────────────

    def _extract_rows(self) -> list[dict]:
        """Normalize various report data structures into rows."""
        rows = []

        # Financial reports with accounts/entries lists
        for key in ['accounts', 'entries', 'items', 'rows', 'income', 'expenses', 'assets', 'liabilities', 'equity']:
            items = self.data.get(key, [])
            if items and isinstance(items, list):
                for item in items:
                    if isinstance(item, dict):
                        # Flatten balance fields
                        if 'balance' in item:
                            item['balance'] = round(float(item['balance']), 2)
                        if 'total_debit' in item:
                            item['total_debit'] = round(float(item['total_debit']), 2)
                        if 'total_credit' in item:
                            item['total_credit'] = round(float(item['total_credit']), 2)
                        rows.append(item)
                if rows:
                    return rows

        # Aging report with bucket items
        buckets = self.data.get('buckets', {})
        if buckets:
            for bucket_name, bucket_data in buckets.items():
                for item in bucket_data.get('items', []):
                    item['bucket'] = bucket_name
                    rows.append(item)
            if rows:
                return rows

        # GST report with input/output sections
        for section in ['output_gst', 'input_gst_itc']:
            sdata = self.data.get(section, {})
            if sdata:
                sdata['_section'] = section
                rows.append(sdata)

        return rows


class ExportManager:
    """Manages export jobs, history, and file storage."""

    def __init__(self, storage_path: str = "/tmp/exports"):
        self.storage_path = storage_path
        os.makedirs(storage_path, exist_ok=True)

    def create_export(self, report_data: dict, report_name: str, fmt: str) -> dict:
        """Generate export file and return download info."""
        exporter = ReportExporter(report_data, report_name)
        export_id = str(uuid.uuid4())

        if fmt == ExportFormat.XLSX:
            content = exporter.to_xlsx()
            ext = 'xlsx'
            mime = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        elif fmt == ExportFormat.CSV:
            content = exporter.to_csv()
            ext = 'csv'
            mime = 'text/csv'
        else:
            content = exporter.to_json()
            ext = 'json'
            mime = 'application/json'

        filename = f"{report_name.replace(' ', '_').lower()}_{export_id[:8]}.{ext}"
        filepath = os.path.join(self.storage_path, filename)
        with open(filepath, 'wb') as f:
            f.write(content)

        return {
            'export_id': export_id,
            'filename': filename,
            'file_size': len(content),
            'mime_type': mime,
            'format': fmt,
            'download_url': f'/api/v1/reports/exports/{export_id}/download',
        }
